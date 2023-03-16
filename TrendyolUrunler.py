import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

url = "https://www.trendyol.com/laptop-x-c103108"
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36"
}
response = requests.get(url, headers=headers)
soup = BeautifulSoup(response.text, "html.parser")

products = soup.find_all("div", class_="p-card-wrppr")

product_data = []

for product in products:
    product_dict = {}
    product_link = product.find("a")["href"]
    product_dict["product_link"] = f"https://www.trendyol.com{product_link}"
    
    product_name = product.find("span", class_="prdct-desc-cntnr-name")
    if product_name:
        product_dict["product_name"] = product_name["title"]
        
    product_price = product.find("div", class_="prc-box-dscntd")
    if product_price:
        product_dict["product_price"] = product_price.text.strip().replace(" TL", "").replace(".", "").replace(",", ".")
        
    product_rating = product.find("div", class_="full")
    if product_rating:
        rating_width = product_rating["style"].split(";")[0]
        product_dict["product_rating"] = float(rating_width.split(":")[1].split("%")[0]) / 20
    else:
        product_dict["product_rating"] = None
        
    # Ürün resimlerini al
    product_image = product.find("img", class_="p-card-img")
    if product_image:
        product_image_src = product_image.get("src")
    else:
        product_image_src = "No image found"
    product_dict["product_image"] = product_image_src
    product_data.append(product_dict)

df = pd.DataFrame(product_data)

current_time = datetime.now()
formatted_time = current_time.strftime("%Y-%m-%d_%H-%M-%S")
file_name = f"laptops_{formatted_time}.xlsx"

# Excel dosyasını oluşturma ve köprüleri ekleme
wb = openpyxl.Workbook()
ws = wb.active

for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

for cell in ws["A"] + ws["1"]:
    cell.style = "Pandas"

for row in ws.iter_rows(min_row=2, max_col=1, max_row=len(product_data) + 1):
    for cell in row:
        cell.value = f'=HYPERLINK("{cell.value}", "Ürün Link")'

for row in ws.iter_rows(min_row=2, min_col=5, max_row=len(product_data) + 1, max_col=5):
    for cell in row:
        cell.value = f'=HYPERLINK("{cell.value}", "Resmi Gör")'

# Dosyayı kaydet
wb.save(file_name)
